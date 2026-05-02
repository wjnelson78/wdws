#!/usr/bin/env python3
"""
ACP Attachment Ingest Pipeline v1.0

Extracts attachment binaries from .eml source files on disk,
runs full page-by-page OCR with parallel workers (same quality as ingest.py),
then chunks → embeds → stores as core.documents + core.document_chunks,
and links back to legal.email_attachments.

This replaces the old email_agent.py approach which used basic text extraction
that missed content. This does proper 150 DPI OCR page-by-page.

Usage:
    python ingest_attachments.py                    # Process all unlinked attachments
    python ingest_attachments.py --workers 10       # 10 OCR workers
    python ingest_attachments.py --re-ocr           # Re-OCR even if extracted_text exists
    python ingest_attachments.py --dry-run           # Just count, don't process
"""

import argparse
import asyncio
from concurrent.futures import ProcessPoolExecutor, as_completed
import email as email_lib
from email import policy
import hashlib
import json
import logging
import multiprocessing
import os
import re
import shutil
import subprocess
import sys
sys.path.insert(0, "/opt/wdws")
import tempfile
import time
import uuid
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import asyncpg
import httpx
from PIL import Image

from ingestion_pipeline import (
    DocumentSpec,
    ingest as pipeline_ingest,
)
from embedding_service import (
    embed_texts_sync, embed_query_sync, _vec_literal as _embedding_vec_literal,
    EMBEDDING_DIMENSIONS, EMBEDDING_MODEL,
)
from contextual_retrieval import generate_context_sync, enrich_chunks

# ============================================================
# Configuration
# ============================================================

# Load .env
from pathlib import Path as _Path
_env_file = _Path("/opt/wdws/.env")
if _env_file.exists():
    for _line in _env_file.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            import os as _os
            _os.environ.setdefault(_k.strip(), _v.strip())

DATABASE_URL = os.environ["DATABASE_URL"]
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
# EMBEDDING_MODEL and EMBEDDING_DIMENSIONS imported from embedding_service
# (BGE-M3 local model, 1024 dimensions)

EMAILS_DIR = Path(os.getenv("EMAILS_DIR", "/opt/wdws/data/emails"))
EXTRACT_DIR = Path("/opt/wdws/data/extracted_attachments")

CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200
OCR_DPI = 150
EMBEDDING_BATCH_SIZE = 100

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("att-ingest")


# ============================================================
# Case number extraction (same as ingest.py)
# ============================================================

STATE_CASE_RE = re.compile(r'(\d{2})-(\d)-(\d{5})-(\d{2})')
FEDERAL_CASE_RE = re.compile(
    r'(\d{1,2})[-:]?(\d{2})[-:]?cv[-:]?(\d{5})[-:]?(\w{2,4})?', re.IGNORECASE
)

def extract_case_numbers_from_text(text: str) -> List[str]:
    found = set()
    for m in FEDERAL_CASE_RE.finditer(text):
        court, year, seq, judge = m.groups()
        normalized = f"{court}{year}CV{seq}"
        if judge:
            normalized += judge.upper()
        found.add(normalized.upper())
    for m in STATE_CASE_RE.finditer(text):
        found.add(m.group(0))
    return list(found)


# ============================================================
# Text chunker (same as ingest.py)
# ============================================================

class TextChunker:
    def __init__(self, chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.separators = ["\n\n", "\n", ". ", " ", ""]

    def split(self, text: str) -> List[str]:
        if not text:
            return []
        return self._split_text(text, self.separators)

    def _split_text(self, text: str, separators: list) -> list:
        final_chunks = []
        separator = separators[-1]
        new_separators = []
        for i, sep in enumerate(separators):
            if sep == "" or sep in text:
                separator = sep
                new_separators = separators[i + 1:]
                break
        splits = text.split(separator) if separator else list(text)
        good_splits = []
        for s in splits:
            if len(s) < self.chunk_size:
                good_splits.append(s)
            else:
                if good_splits:
                    merged = self._merge_splits(good_splits, separator)
                    final_chunks.extend(merged)
                    good_splits = []
                if new_separators:
                    final_chunks.extend(self._split_text(s, new_separators))
                else:
                    final_chunks.append(s)
        if good_splits:
            merged = self._merge_splits(good_splits, separator)
            final_chunks.extend(merged)
        return final_chunks

    def _merge_splits(self, splits: list, separator: str) -> list:
        docs = []
        current = []
        total = 0
        for s in splits:
            s_len = len(s)
            if total + s_len + (len(separator) if current else 0) > self.chunk_size:
                if current:
                    docs.append(separator.join(current))
                    while total > self.chunk_overlap and len(current) > 1:
                        total -= len(current[0]) + len(separator)
                        current.pop(0)
                current = [s]
                total = s_len
            else:
                current.append(s)
                total += s_len + (len(separator) if len(current) > 1 else 0)
        if current:
            docs.append(separator.join(current))
        return docs


# ============================================================
# OCR Functions (same as ingest.py — full quality)
# ============================================================

def extract_pdf_text(pdf_path: Path) -> str:
    """Extract text from PDF using OCR — page-by-page to limit memory."""
    import pytesseract
    from pdf2image import convert_from_path, pdfinfo_from_path

    try:
        try:
            info = pdfinfo_from_path(str(pdf_path))
            num_pages = info.get("Pages", 0)
        except Exception:
            num_pages = 0

        if num_pages > 0:
            parts = []
            for page_num in range(1, num_pages + 1):
                try:
                    images = convert_from_path(
                        str(pdf_path), dpi=OCR_DPI,
                        first_page=page_num, last_page=page_num,
                        thread_count=1
                    )
                    if images:
                        page_text = pytesseract.image_to_string(images[0])
                        if page_text and page_text.strip():
                            parts.append(f"[Page {page_num}]\n{page_text}")
                        del images
                except Exception as e:
                    log.debug(f"  OCR page {page_num}/{num_pages} failed: {e}")
            return "\n\n".join(parts)
        else:
            images = convert_from_path(str(pdf_path), dpi=OCR_DPI, thread_count=1)
            parts = []
            for page_num, img in enumerate(images, 1):
                page_text = pytesseract.image_to_string(img)
                if page_text and page_text.strip():
                    parts.append(f"[Page {page_num}]\n{page_text}")
            return "\n\n".join(parts)
    except Exception as e:
        log.warning(f"OCR failed for {pdf_path.name}: {e}, trying pdftotext")
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", str(pdf_path), "-"],
                capture_output=True, text=True, timeout=30
            )
            return result.stdout
        except Exception as e2:
            log.error(f"All extraction failed for {pdf_path.name}: {e2}")
            return ""


def extract_image_text(image_path: Path) -> str:
    """Extract text from image using OCR."""
    import pytesseract
    try:
        img = Image.open(image_path)
        return pytesseract.image_to_string(img)
    except Exception as e:
        log.error(f"Image OCR failed for {image_path.name}: {e}")
        return ""


def extract_docx_text(docx_path: Path) -> str:
    """Extract text from DOCX file."""
    try:
        with zipfile.ZipFile(docx_path) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                paragraphs = root.findall(".//w:p", ns)
                texts = []
                for p in paragraphs:
                    runs = p.findall(".//w:t", ns)
                    line = "".join(r.text or "" for r in runs)
                    if line:
                        texts.append(line)
                return "\n".join(texts)
    except Exception as e:
        log.error(f"DOCX extraction failed for {docx_path.name}: {e}")
        return ""


def extract_xlsx_text(xlsx_path: Path) -> str:
    """Extract text from XLSX via openpyxl. Sheet-prefixed TSV; skips blank rows."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        parts: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [("" if c is None else str(c)) for c in row]
                if any(c.strip() for c in cells):
                    parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)
    except Exception as e:
        log.error(f"XLSX extraction failed for {xlsx_path.name}: {e}")
        return ""


def extract_xls_text(xls_path: Path) -> str:
    """Extract text from legacy XLS via soffice → xlsx → openpyxl."""
    try:
        with tempfile.TemporaryDirectory() as td:
            result = subprocess.run(
                ["soffice", "--headless", "--convert-to", "xlsx",
                 "--outdir", td, str(xls_path)],
                capture_output=True, timeout=120,
            )
            if result.returncode != 0:
                log.error(
                    f"soffice xls→xlsx failed for {xls_path.name}: "
                    f"{result.stderr.decode(errors='ignore')[:200]}"
                )
                return ""
            converted = next(Path(td).glob("*.xlsx"), None)
            if not converted:
                return ""
            return extract_xlsx_text(converted)
    except Exception as e:
        log.error(f"XLS extraction failed for {xls_path.name}: {e}")
        return ""


# ============================================================
# Attachment extraction from .eml files
# ============================================================

def extract_attachments_from_eml(eml_path: Path) -> List[Tuple[str, str, bytes]]:
    """Extract all attachments from an .eml file.
    Returns list of (filename, content_type, binary_data)."""
    try:
        with open(eml_path, "rb") as f:
            msg = email_lib.message_from_binary_file(f, policy=policy.default)

        attachments = []
        if msg.is_multipart():
            for part in msg.walk():
                disp = str(part.get("Content-Disposition", ""))
                fname = part.get_filename()

                # It's an attachment if it has a filename or attachment disposition
                if not fname and "attachment" not in disp.lower():
                    continue
                if not fname:
                    ct = part.get_content_type()
                    ext = {
                        "application/pdf": ".pdf",
                        "image/png": ".png",
                        "image/jpeg": ".jpg",
                        "text/plain": ".txt",
                    }.get(ct, ".bin")
                    fname = f"attachment{ext}"

                payload = part.get_payload(decode=True)
                if payload:
                    ct = part.get_content_type() or "application/octet-stream"
                    attachments.append((fname, ct, payload))

        return attachments
    except Exception as e:
        log.error(f"Failed to parse {eml_path}: {e}")
        return []


# ============================================================
# OCR worker (top-level for multiprocessing pickle)
# ============================================================

RPMSG_MAGIC = bytes.fromhex("76e80460c411e386a00f000000100000")


def _is_rpmsg(file_path: Path, content_type: str) -> bool:
    """Detect Microsoft RMS-protected email wrappers (.rpmsg / message_v2.rpmsg)."""
    if content_type == "application/x-microsoft-rpmsg-message":
        return True
    if file_path.suffix.lower() == ".rpmsg":
        return True
    try:
        if file_path.stat().st_size >= 16:
            with open(file_path, "rb") as fh:
                return fh.read(16) == RPMSG_MAGIC
    except Exception:
        pass
    return False


def _extract_rpmsg_text(file_path: Path) -> Tuple[str, str]:
    """Decrypt a .rpmsg via MIP SDK and return (concatenated_text, method).

    The decrypted body + any inner attachments (AIP-protected PDFs get
    transparently double-decrypted by purview_decrypt's GetDecryptedTemporaryFileAsync
    path) are concatenated into one text blob suitable for chunk+embed.
    See PURVIEW_DECRYPT.md for architecture details.
    """
    try:
        import sys as _sys
        _sys.path.insert(0, "/opt/wdws")
        from purview_decrypt import decrypt_rpmsg
    except Exception as e:
        return "", f"error: rpmsg_import_failed: {e}"

    # Write decrypted artifacts next to the .rpmsg
    body_out = file_path.with_suffix("").with_name(file_path.stem + ".decrypted.html")
    result = decrypt_rpmsg(str(file_path), str(body_out), timeout=120)
    status = result.get("status")
    if status == "auth_required":
        return "", f"rpmsg_auth_required:{result.get('required_tenant','?')}"
    if status == "no_permissions":
        return "", f"rpmsg_no_permissions:pl_owner={result.get('plOwner','?')}"
    if status != "ok":
        return "", f"rpmsg_{status}:{result.get('error','')[:120]}"

    # Strip HTML body → text
    parts: List[str] = []
    parts.append(f"[RMS-protected email; pl_owner={result.get('plOwner')}; bodyType={result.get('bodyType')}]")
    try:
        body_raw = body_out.read_text(errors="replace")
        body_stripped = re.sub(r"<style.*?</style>", "", body_raw, flags=re.DOTALL | re.I)
        body_stripped = re.sub(r"<script.*?</script>", "", body_stripped, flags=re.DOTALL | re.I)
        body_stripped = re.sub(r"<[^>]+>", " ", body_stripped)
        import html as _html
        body_stripped = _html.unescape(body_stripped)
        body_stripped = re.sub(r"\s+", " ", body_stripped).strip()
        parts.append(body_stripped)
    except Exception as e:
        parts.append(f"[body extract error: {e}]")

    # Inner attachments — purview_decrypt already wrote them as .attNN_* siblings
    for att_meta in result.get("attachments", []):
        att_path = Path(att_meta["output"])
        if not att_path.exists():
            continue
        ext = att_path.suffix.lower()
        att_text = ""
        try:
            if ext == ".pdf":
                # purview_decrypt's File API wrote decrypted PDF bytes on write.
                # In practice, inner PDFs are often AIP-protected themselves,
                # leaving only a "This is a protected document" cover page.
                # The fix: run a second decrypt pass via purview-decrypt on the PDF.
                dec = att_path.with_suffix(".inner_dec.pdf")
                inner = decrypt_rpmsg(str(att_path), str(dec), timeout=120)
                pdf_to_read = dec if (inner.get("status") == "ok" and dec.exists()) else att_path
                att_text = extract_pdf_text(pdf_to_read)
            elif ext in (".png", ".jpg", ".jpeg", ".gif"):
                # Skip tiny inline signatures/logos
                if att_path.stat().st_size >= 30_000:
                    att_text = extract_image_text(att_path)
            elif ext in (".docx",):
                att_text = extract_docx_text(att_path)
            elif ext == ".txt":
                att_text = att_path.read_text(errors="ignore")
        except Exception as e:
            att_text = f"[attachment extract error: {e}]"
        if att_text:
            parts.append(f"\n[attachment: {att_meta.get('name')}]\n{att_text}")

    return "\n\n".join(parts), "rpmsg_decrypt"


def _ocr_attachment_worker(args: Tuple[str, str, str]) -> Tuple[str, str, str]:
    """Worker: OCR a single attachment file.
    Args: (attachment_db_id, file_path, content_type)
    Returns: (attachment_db_id, extracted_text, method)
    """
    att_id, file_path_str, content_type = args
    file_path = Path(file_path_str)
    start = time.time()

    try:
        if _is_rpmsg(file_path, content_type):
            text, method = _extract_rpmsg_text(file_path)
        elif content_type == "application/pdf":
            text = extract_pdf_text(file_path)
            method = "ocr_pdf_150dpi"
        elif content_type.startswith("image/"):
            text = extract_image_text(file_path)
            method = "ocr_image"
        elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            text = extract_docx_text(file_path)
            method = "docx_xml"
        elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            text = extract_xlsx_text(file_path)
            method = "xlsx_openpyxl"
        elif content_type == "application/vnd.ms-excel":
            text = extract_xls_text(file_path)
            method = "xls_soffice"
        elif content_type == "text/plain":
            text = file_path.read_text(errors="ignore")
            method = "text_direct"
        else:
            # Try as PDF first, then image
            text = extract_pdf_text(file_path)
            if not text.strip():
                text = extract_image_text(file_path)
                method = "ocr_image_fallback"
            else:
                method = "ocr_pdf_fallback"

        elapsed = time.time() - start
        text_len = len(text.strip()) if text else 0
        log.info(f"  ⚙ OCR done: {file_path.name} ({content_type}) → {text_len} chars in {elapsed:.1f}s")
        return (att_id, text.strip() if text else "", method)
    except Exception as e:
        elapsed = time.time() - start
        log.error(f"  ✗ OCR failed: {file_path.name}: {e} ({elapsed:.1f}s)")
        return (att_id, "", f"error: {str(e)[:100]}")


# ============================================================
# Embedding client (same as ingest.py)
# ============================================================

class EmbeddingClient:
    """Local BGE-M3 embedding client with batching (delegates to embedding_service)."""

    def __init__(self, **kwargs):
        self.total_tokens = 0
        self.total_requests = 0

    async def embed_batch(self, texts: List[str]) -> List[List[float]]:
        if not texts:
            return []
        result = embed_texts_sync(texts)
        self.total_requests += 1
        return result

    async def close(self):
        pass  # No HTTP client to close


# ============================================================
# Progress tracker
# ============================================================

class ProgressTracker:
    def __init__(self, total: int, label: str):
        self.total = total
        self.label = label
        self.done = 0
        self.ok = 0
        self.skipped = 0
        self.errors = 0
        self.chunks_written = 0
        self.start_time = time.time()

    def tick(self, success: bool = True, chunks: int = 0, filename: str = "", detail: str = ""):
        self.done += 1
        if success:
            self.ok += 1
            self.chunks_written += chunks
        elif detail == "skipped":
            self.skipped += 1
        else:
            self.errors += 1

        elapsed = time.time() - self.start_time
        rate = self.done / elapsed * 60 if elapsed > 0 else 0
        eta = (self.total - self.done) / (self.done / elapsed) if self.done > 0 and elapsed > 0 else 0
        pct = (self.done / self.total * 100) if self.total > 0 else 100
        icon = "✓" if success else ("⊘" if detail == "skipped" else "✗")
        msg = f"  [{self.done}/{self.total} {pct:.0f}%] {icon} {filename}"
        if chunks:
            msg += f" → {chunks} chunks"
        if detail and detail != "skipped":
            msg += f" ({detail})"
        eta_str = self._fmt_time(eta)
        rate_str = f"{rate:.1f}"
        msg += f"  [{rate_str}/min, ETA {eta_str}]"
        log.info(msg)

    def summary(self) -> str:
        elapsed = time.time() - self.start_time
        rate = self.ok / (elapsed / 60) if elapsed > 0 else 0
        return (
            f"{self.label}: {self.ok} ok, {self.skipped} skipped, {self.errors} errors, "
            f"{self.chunks_written} chunks in {self._fmt_time(elapsed)} ({rate:.1f} docs/min)"
        )

    @staticmethod
    def _fmt_time(seconds: float) -> str:
        if seconds < 60:
            return f"{seconds:.0f}s"
        elif seconds < 3600:
            return f"{seconds/60:.1f}m"
        else:
            h = int(seconds // 3600)
            m = int((seconds % 3600) // 60)
            return f"{h}h{m:02d}m"


# ============================================================
# Database helpers
# ============================================================

def _vec_literal(embedding: List[float]) -> str:
    return _embedding_vec_literal(embedding)


async def get_unlinked_attachments(pool: asyncpg.Pool) -> List[dict]:
    """Get all attachments that don't have a child_doc_id yet."""
    rows = await pool.fetch("""
        SELECT
            ea.id, ea.email_doc_id, ea.filename, ea.content_type, ea.file_size,
            d.source_path AS email_source_path,
            COALESCE(em.sender, '') AS email_sender,
            COALESCE(d.title, '') AS email_subject
        FROM legal.email_attachments ea
        JOIN core.documents d ON d.id = ea.email_doc_id
        LEFT JOIN legal.email_metadata em ON em.document_id = ea.email_doc_id
        WHERE ea.child_doc_id IS NULL
        ORDER BY ea.created_at
    """)
    return [dict(r) for r in rows]




# ============================================================
# Main pipeline
# ============================================================

async def main():
    parser = argparse.ArgumentParser(description="ACP Attachment Ingest Pipeline")
    parser.add_argument("--workers", type=int, default=10, help="OCR parallel workers")
    parser.add_argument("--re-ocr", action="store_true", help="Re-OCR all, even if already linked")
    parser.add_argument("--dry-run", action="store_true", help="Count only, don't process")
    parser.add_argument("--limit", type=int, default=0, help="Process only N attachments (0=all)")
    args = parser.parse_args()

    log.info(f"Connecting to PostgreSQL...")
    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)

    # Get unlinked attachments from DB
    if args.re_ocr:
        all_atts = await pool.fetch("""
            SELECT ea.id, ea.email_doc_id, ea.filename, ea.content_type, ea.file_size,
                   d.source_path AS email_source_path,
                   COALESCE(d.title, '') AS email_subject
            FROM legal.email_attachments ea
            JOIN core.documents d ON d.id = ea.email_doc_id
            ORDER BY ea.created_at
        """)
        atts = [dict(r) for r in all_atts]
    else:
        atts = await get_unlinked_attachments(pool)

    if args.limit > 0:
        atts = atts[:args.limit]

    total = len(atts)
    log.info(f"")
    log.info(f"╔{'═'*58}╗")
    log.info(f"║  ACP Attachment Ingest Pipeline v1.0{' '*21}║")
    log.info(f"╠{'═'*58}╣")
    log.info(f"║  Attachments to process: {total:<33}║")
    log.info(f"║  OCR workers:           {args.workers:<34}║")
    log.info(f"║  OCR DPI:               {OCR_DPI:<34}║")
    log.info(f"║  Re-OCR mode:           {str(args.re_ocr):<34}║")
    log.info(f"║  Emails dir:            {str(EMAILS_DIR)[:34]:<34}║")
    log.info(f"╚{'═'*58}╝")
    log.info(f"")

    if total == 0:
        log.info("No unlinked attachments found. Nothing to do.")
        await pool.close()
        return

    if args.dry_run:
        by_type = {}
        for a in atts:
            ct = a["content_type"] or "unknown"
            by_type[ct] = by_type.get(ct, 0) + 1
        log.info(f"Would process {total} attachments:")
        for ct, cnt in sorted(by_type.items(), key=lambda x: -x[1]):
            log.info(f"  {ct}: {cnt}")
        await pool.close()
        return

    # ── Phase 1: Extract attachment binaries from .eml files ──
    log.info(f"{'─'*60}")
    log.info(f"  Phase 1: Extracting attachment binaries from .eml files")
    log.info(f"{'─'*60}")

    EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
    extracted: List[Tuple[str, str, str]] = []  # (att_id, file_path, content_type)
    extract_errors = 0

    # Group attachments by source .eml file to avoid re-parsing
    by_eml: Dict[str, List[dict]] = {}
    for att in atts:
        src = att["email_source_path"]
        if src not in by_eml:
            by_eml[src] = []
        by_eml[src].append(att)

    log.info(f"  {total} attachments across {len(by_eml)} unique .eml files")

    for eml_path_str, att_list in by_eml.items():
        eml_path = Path(eml_path_str)
        if not eml_path.exists():
            log.warning(f"  ⊘ .eml not found: {eml_path_str}")
            extract_errors += len(att_list)
            continue

        # Extract all attachments from this .eml
        eml_attachments = extract_attachments_from_eml(eml_path)

        if not eml_attachments:
            log.debug(f"  ⊘ No extractable attachments in {eml_path.name}")
            extract_errors += len(att_list)
            continue

        # Match DB records to extracted files by filename
        for att in att_list:
            db_filename = att["filename"]
            db_content_type = att["content_type"] or ""

            # Find best matching attachment
            matched = None
            for fname, ct, data in eml_attachments:
                if fname == db_filename:
                    matched = (fname, ct, data)
                    break
            # Fallback: match by content type if filename didn't match
            if not matched:
                for fname, ct, data in eml_attachments:
                    if ct == db_content_type:
                        matched = (fname, ct, data)
                        break
            # Last resort: just take the first unmatched one
            if not matched and eml_attachments:
                matched = eml_attachments[0]

            if matched:
                fname, ct, data = matched
                # Save to disk
                safe_name = re.sub(r'[^\w\-\.]', '_', fname)
                att_id_str = str(att["id"])
                save_path = EXTRACT_DIR / f"{att_id_str}_{safe_name}"
                save_path.write_bytes(data)
                extracted.append((att_id_str, str(save_path), ct))
                # Remove from list to avoid double-matching
                eml_attachments.remove(matched)
            else:
                log.warning(f"  ⊘ Could not match attachment '{db_filename}' in {eml_path.name}")
                extract_errors += 1

    log.info(f"  ✔ Extracted {len(extracted)} files to disk ({extract_errors} errors)")

    if not extracted:
        log.error("No attachments extracted. Exiting.")
        await pool.close()
        return

    # ── Phase 2: Parallel OCR ──
    log.info(f"")
    log.info(f"{'─'*60}")
    log.info(f"  Phase 2: OCR with {args.workers} parallel workers @ {OCR_DPI} DPI")
    log.info(f"{'─'*60}")

    ocr_results: Dict[str, Tuple[str, str]] = {}  # att_id → (text, method)
    ocr_tracker = ProgressTracker(len(extracted), "OCR")

    with ProcessPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(_ocr_attachment_worker, item): item for item in extracted}
        for future in as_completed(futures):
            item = futures[future]
            att_id = item[0]
            try:
                att_id_out, text, method = future.result()
                if text and len(text) >= 20:
                    ocr_results[att_id_out] = (text, method)
                    ocr_tracker.tick(success=True, filename=Path(item[1]).name)
                else:
                    ocr_tracker.tick(success=False, filename=Path(item[1]).name, detail="no text")
            except Exception as e:
                log.error(f"  ✗ OCR exception for {Path(item[1]).name}: {e}")
                ocr_tracker.tick(success=False, filename=Path(item[1]).name, detail=str(e)[:60])

    log.info(f"  ✔ {ocr_tracker.summary()}")

    # ── Phase 3: Chunk + Embed + Write ──
    log.info(f"")
    log.info(f"{'─'*60}")
    log.info(f"  Phase 3: Chunk → Embed → Store ({len(ocr_results)} documents)")
    log.info(f"{'─'*60}")

    # Build a lookup from att_id → att metadata
    att_lookup = {str(a["id"]): a for a in atts}

    chunker = TextChunker()
    embedder = EmbeddingClient()
    write_tracker = ProgressTracker(len(ocr_results), "Embed+Write")
    total_chunks = 0
    total_docs = 0
    write_errors = 0

    for att_id, (text, method) in ocr_results.items():
        att_info = att_lookup.get(att_id)
        if not att_info:
            log.error(f"  ✗ No metadata for attachment {att_id}")
            write_errors += 1
            write_tracker.tick(success=False, filename="unknown", detail="no metadata")
            continue

        filename = att_info["filename"]
        email_subject = att_info.get("email_subject", "")
        email_doc_id = att_info["email_doc_id"]
        email_source_path = att_info["email_source_path"]

        # Compute title first so the pipeline (and downstream logs) have it.
        # Note: the previous version referenced `title` before assigning it —
        # that was a latent bug (only survived because the first att_id
        # iteration was always the re-entry of a process restart). Fixed here.
        title = f"Attachment: {filename}"
        if email_subject:
            title = f"{filename} (from: {email_subject[:80]})"

        # Hand off to the unified pipeline — handles chunking, Contextual
        # Retrieval, embedding, core.documents / core.document_chunks /
        # legal.case_documents / core.document_relationships (parent→child)
        # and the legal.email_attachments upsert via attachment_meta.
        # Auto-extracts case numbers from the OCR'd text.
        #
        # source_path must be UNIQUE per attachment — the pipeline uses it
        # as its idempotency key. Previously the old write path reused the
        # parent email's source_path across all attachments; under the
        # pipeline that would cause the second attachment to match the first
        # and get dedup'd as was_existing. Disambiguate with the att_id.
        att_source_path = f"{email_source_path}#attachment/{att_id}"
        spec = DocumentSpec(
            domain="legal",
            document_type="email_attachment",
            source_path=att_source_path,
            title=title,
            filename=filename,
            full_content=text,
            parent_doc_id=str(email_doc_id),
            relationship_type="has_attachment",
            metadata={
                "parent_email_doc_id": str(email_doc_id),
                "parent_email_source_path": email_source_path,
                "attachment_id": str(att_id),
                "extraction_method": method,
                "source": "ingest_attachments",
            },
            attachment_meta={
                "id": str(att_id),
                "filename": filename,
                "extracted_text": text,
                "extraction_method": method,
            },
        )

        try:
            result = await pipeline_ingest(spec, pool=pool)
        except Exception as e:
            log.error(f"  ✗ Ingest failed for {filename}: {e}")
            write_errors += 1
            write_tracker.tick(success=False, filename=filename, detail=str(e)[:60])
            continue

        if result.error:
            log.error(f"  ✗ Ingest error for {filename}: {result.error}")
            write_errors += 1
            write_tracker.tick(success=False, filename=filename, detail=result.error[:60])
            continue

        if result.chunk_count == 0:
            write_tracker.tick(success=False, filename=filename, detail="no chunks")
            continue

        total_docs += 1
        total_chunks += result.chunk_count
        write_tracker.tick(
            success=True,
            chunks=result.chunk_count,
            filename=filename,
        )

    log.info(f"  ✔ {write_tracker.summary()}")

    # ── Cleanup extracted files ──
    log.info(f"Cleaning up extracted files...")
    try:
        shutil.rmtree(EXTRACT_DIR)
    except Exception:
        pass

    # ── Final stats ──
    final_docs = await pool.fetchval("SELECT COUNT(*) FROM core.documents")
    final_chunks = await pool.fetchval("SELECT COUNT(*) FROM core.document_chunks")
    linked = await pool.fetchval("SELECT COUNT(*) FROM legal.email_attachments WHERE child_doc_id IS NOT NULL")
    embed_cost = embedder.total_tokens / 1_000_000 * 0.13

    await embedder.close()
    await pool.close()

    log.info(f"")
    log.info(f"╔{'═'*58}╗")
    log.info(f"║  ATTACHMENT INGEST COMPLETE{' '*31}║")
    log.info(f"╠{'═'*58}╣")
    log.info(f"║  Extracted from .eml:   {len(extracted):<34}║")
    log.info(f"║  OCR successful:        {ocr_tracker.ok:<34}║")
    log.info(f"║  Documents created:     {total_docs:<34}║")
    log.info(f"║  Chunks created:        {total_chunks:<34}║")
    log.info(f"║  Extraction errors:     {extract_errors:<34}║")
    log.info(f"║  OCR errors/empty:      {ocr_tracker.errors:<34}║")
    log.info(f"║  Write errors:          {write_errors:<34}║")
    log.info(f"╠{'═'*58}╣")
    log.info(f"║  DATABASE TOTALS{' '*41}║")
    log.info(f"║  Total documents:       {final_docs:<34}║")
    log.info(f"║  Total chunks:          {final_chunks:<34}║")
    log.info(f"║  Linked attachments:    {linked:<34}║")
    log.info(f"╠{'═'*58}╣")
    log.info(f"║  EMBEDDING COST{' '*42}║")
    log.info(f"║  Tokens:                {embedder.total_tokens:,}{' '*(34-len(f'{embedder.total_tokens:,}'))}║")
    log.info(f"║  Requests:              {embedder.total_requests:<34}║")
    log.info(f"║  Est. cost:             ${embed_cost:.4f}{' '*(33-len(f'${embed_cost:.4f}'))}║")
    log.info(f"╚{'═'*58}╝")


if __name__ == "__main__":
    asyncio.run(main())
