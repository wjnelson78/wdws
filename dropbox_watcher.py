#!/usr/bin/env python3
"""
WDWS Dropbox Watcher — Samba Share Auto-Ingestion

Watches /opt/wdws/data/dropbox/{medical,legal,general} for new files dropped
via Samba share. Automatically:
  1. Detects new files (PDF, DOCX, images, CCDA XML, TXT, EML)
  2. Extracts text via OCR / parsing
  3. Chunks text
  4. Generates embeddings (OpenAI text-embedding-3-large)
  5. Stores in PostgreSQL (core.documents + core.document_chunks)
  6. Enriches medical/legal metadata tables
  7. Moves processed files to /opt/wdws/data/dropbox/.processed/

Drop files into subdirectories:
  medical/  → ingested as medical domain (UW records, PDFs, images, CCDA)
  legal/    → ingested as legal domain (court filings, emails)
  general/  → ingested as general domain (miscellaneous documents)

Usage:
    python dropbox_watcher.py              # Run as daemon
    python dropbox_watcher.py --once       # Process pending files and exit
"""

import asyncio
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys
sys.path.insert(0, "/opt/wdws")
import time
import uuid
import xml.etree.ElementTree as ET
import zipfile
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import asyncpg
import httpx
from inotify_simple import INotify, flags as inotify_flags
from PIL import Image

from embedding_service import (
    embed_texts_sync, embed_query_sync,
    _vec_literal as _embedding_vec_literal,
    EMBEDDING_DIMENSIONS, EMBEDDING_MODEL,
)
from contextual_retrieval import generate_context_sync, enrich_chunks

# ============================================================
# Configuration
# ============================================================

DATABASE_URL = os.environ["DATABASE_URL"]
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
# EMBEDDING_MODEL and EMBEDDING_DIMENSIONS imported from embedding_service
# (BGE-M3 local model, 1024 dimensions)

DROPBOX_DIR = Path("/opt/wdws/data/dropbox")
PROCESSED_DIR = DROPBOX_DIR / ".processed"
FAILED_DIR = DROPBOX_DIR / ".failed"

CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200
OCR_DPI = 150

# Supported file extensions by domain
SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".txt", ".eml",
    ".png", ".jpg", ".jpeg", ".tiff", ".tif",
    ".xml", ".csv", ".rtf", ".xlsx", ".xls",
}

# Settle time — wait this long after last write before processing
# (Samba writes can come in chunks)
SETTLE_SECONDS = 5

try:
    from athena_logging import setup_logging, get_logger
    setup_logging("dropbox-watcher")
    log = get_logger()
except ImportError:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s [%(name)s] %(message)s", datefmt="%H:%M:%S")
    log = logging.getLogger("dropbox-watcher")


# ============================================================
# Text extraction (reused from ingest.py)
# ============================================================

def extract_pdf_text(pdf_path: Path) -> str:
    """Extract text from PDF using OCR — page-by-page."""
    try:
        from pdf2image import convert_from_path, pdfinfo_from_path
        import pytesseract

        try:
            info = pdfinfo_from_path(str(pdf_path))
            num_pages = info.get("Pages", 0)
        except Exception:
            num_pages = 0

        if num_pages > 0:
            parts = []
            for page_num in range(1, num_pages + 1):
                try:
                    images = convert_from_path(
                        str(pdf_path), dpi=OCR_DPI,
                        first_page=page_num, last_page=page_num,
                        thread_count=1
                    )
                    if images:
                        page_text = pytesseract.image_to_string(images[0])
                        if page_text and page_text.strip():
                            parts.append(f"[Page {page_num}]\n{page_text}")
                        del images
                except Exception as e:
                    log.debug(f"OCR page {page_num} failed for {pdf_path.name}: {e}")
            return "\n\n".join(parts)
        else:
            images = convert_from_path(str(pdf_path), dpi=OCR_DPI, thread_count=1)
            parts = []
            for page_num, img in enumerate(images, 1):
                page_text = pytesseract.image_to_string(img)
                if page_text and page_text.strip():
                    parts.append(f"[Page {page_num}]\n{page_text}")
            return "\n\n".join(parts)
    except Exception as e:
        log.warning(f"OCR failed for {pdf_path.name}: {e}, trying pdftotext")
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", str(pdf_path), "-"],
                capture_output=True, text=True, timeout=60
            )
            return result.stdout
        except Exception as e2:
            log.error(f"All extraction failed for {pdf_path.name}: {e2}")
            return ""


def extract_image_text(image_path: Path) -> str:
    """Extract text from image using OCR."""
    try:
        import pytesseract
        img = Image.open(image_path)
        return pytesseract.image_to_string(img)
    except Exception as e:
        log.error(f"Image OCR failed for {image_path.name}: {e}")
        return ""


def extract_docx_text(docx_path: Path) -> str:
    """Extract text from DOCX file."""
    try:
        with zipfile.ZipFile(docx_path) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                paragraphs = root.findall(".//w:p", ns)
                texts = []
                for p in paragraphs:
                    runs = p.findall(".//w:t", ns)
                    line = "".join(r.text or "" for r in runs)
                    if line:
                        texts.append(line)
                return "\n".join(texts)
    except Exception as e:
        log.error(f"DOCX extraction failed for {docx_path.name}: {e}")
        return ""


def extract_txt_text(txt_path: Path) -> str:
    """Read plain text file."""
    try:
        return txt_path.read_text(errors="replace")
    except Exception as e:
        log.error(f"TXT read failed for {txt_path.name}: {e}")
        return ""


def extract_xlsx_text(xlsx_path: Path) -> str:
    """Extract text from XLSX using openpyxl or zipfile fallback."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)
    except ImportError:
        # Fallback: extract shared strings from the XLSX zip
        try:
            with zipfile.ZipFile(xlsx_path) as z:
                if "xl/sharedStrings.xml" in z.namelist():
                    with z.open("xl/sharedStrings.xml") as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                        texts = []
                        for si in root.findall(".//s:t", ns):
                            if si.text:
                                texts.append(si.text)
                        return "\n".join(texts)
            return ""
        except Exception as e:
            log.error(f"XLSX extraction failed for {xlsx_path.name}: {e}")
            return ""
    except Exception as e:
        log.error(f"XLSX extraction failed for {xlsx_path.name}: {e}")
        return ""


def extract_eml_text(eml_path: Path) -> str:
    """Extract text from .eml email file."""
    import email as email_lib
    from email import policy
    try:
        with open(eml_path, "rb") as f:
            msg = email_lib.message_from_binary_file(f, policy=policy.default)

        parts = []
        subject = msg.get("Subject", "")
        sender = msg.get("From", "")
        date = msg.get("Date", "")
        parts.append(f"Subject: {subject}")
        parts.append(f"From: {sender}")
        parts.append(f"Date: {date}")
        parts.append("")

        if msg.is_multipart():
            for part in msg.walk():
                ct = part.get_content_type()
                if ct == "text/plain":
                    body = part.get_content()
                    if isinstance(body, bytes):
                        body = body.decode("utf-8", errors="replace")
                    parts.append(body)
        else:
            body = msg.get_content()
            if isinstance(body, bytes):
                body = body.decode("utf-8", errors="replace")
            parts.append(body)

        return "\n".join(parts)
    except Exception as e:
        log.error(f"EML extraction failed for {eml_path.name}: {e}")
        return ""


def extract_ccda_text(xml_path: Path) -> str:
    """Extract text from CCDA XML medical records."""
    try:
        tree = ET.parse(str(xml_path))
        root = tree.getroot()
        ns = {"hl7": "urn:hl7-org:v3"}

        parts = []

        # Patient name
        for name in root.findall(".//hl7:patient/hl7:name", ns):
            given = name.findtext("hl7:given", "", ns)
            family = name.findtext("hl7:family", "", ns)
            if given or family:
                parts.append(f"Patient: {given} {family}")

        # Document title
        title = root.findtext(".//hl7:title", "", ns)
        if title:
            parts.append(f"Document: {title}")

        # Sections
        for section in root.findall(".//hl7:section", ns):
            sec_title = section.findtext("hl7:title", "", ns)
            if sec_title:
                parts.append(f"\n## {sec_title}")
            text_elem = section.find("hl7:text", ns)
            if text_elem is not None:
                text = ET.tostring(text_elem, encoding="unicode", method="text")
                text = re.sub(r'\s+', ' ', text).strip()
                if text:
                    parts.append(text)

        return "\n".join(parts)
    except Exception as e:
        log.error(f"CCDA extraction failed for {xml_path.name}: {e}")
        return ""


def extract_text(file_path: Path) -> str:
    """Route extraction by file type."""
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf_text(file_path)
    elif ext in (".docx", ".doc"):
        return extract_docx_text(file_path)
    elif ext == ".txt":
        return extract_txt_text(file_path)
    elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".tif"):
        return extract_image_text(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx_text(file_path)
    elif ext == ".eml":
        return extract_eml_text(file_path)
    elif ext == ".xml":
        return extract_ccda_text(file_path)
    elif ext == ".csv":
        return extract_txt_text(file_path)  # CSV is plain text
    elif ext == ".rtf":
        # Try unrtf if available
        try:
            result = subprocess.run(
                ["unrtf", "--text", str(file_path)],
                capture_output=True, text=True, timeout=30
            )
            return result.stdout
        except Exception:
            return extract_txt_text(file_path)
    else:
        log.warning(f"Unsupported file type: {ext} for {file_path.name}")
        return ""


# ============================================================
# Text chunker (same as ingest.py)
# ============================================================

class TextChunker:
    def __init__(self, chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.separators = ["\n\n", "\n", ". ", " ", ""]

    def split(self, text: str) -> List[str]:
        if not text:
            return []
        return self._split_text(text, self.separators)

    def _split_text(self, text: str, separators: list) -> list:
        final_chunks = []
        separator = separators[-1]
        new_separators = []
        for i, sep in enumerate(separators):
            if sep == "" or sep in text:
                separator = sep
                new_separators = separators[i + 1:]
                break
        splits = text.split(separator) if separator else list(text)
        good_splits = []
        for s in splits:
            if len(s) < self.chunk_size:
                good_splits.append(s)
            else:
                if good_splits:
                    merged = self._merge_splits(good_splits, separator)
                    final_chunks.extend(merged)
                    good_splits = []
                if new_separators:
                    final_chunks.extend(self._split_text(s, new_separators))
                else:
                    final_chunks.append(s)
        if good_splits:
            merged = self._merge_splits(good_splits, separator)
            final_chunks.extend(merged)
        return final_chunks

    def _merge_splits(self, splits: list, separator: str) -> list:
        docs = []
        current = []
        total = 0
        for s in splits:
            s_len = len(s)
            if total + s_len + (len(separator) if current else 0) > self.chunk_size:
                if current:
                    docs.append(separator.join(current))
                    while total > self.chunk_overlap and len(current) > 1:
                        total -= len(current[0]) + len(separator)
                        current.pop(0)
                current = [s]
                total = s_len
            else:
                current.append(s)
                total += s_len + (len(separator) if len(current) > 1 else 0)
        if current:
            docs.append(separator.join(current))
        return docs


# ============================================================
# Medical auto-categorization
# ============================================================

MEDICAL_PATTERNS = {
    "lab_result": [r"lab\s*result", r"blood\s*test", r"cbc|cmp|lipid\s*panel",
                   r"hemoglobin|glucose|cholesterol"],
    "radiology": [r"radiol|x-ray|ct\s*scan|mri|ultrasound|imaging"],
    "progress_note": [r"progress\s*note|office\s*visit|follow.up|SOAP\b"],
    "discharge_summary": [r"discharge\s*summar"],
    "operative_report": [r"operat(?:ive|ion)\s*report|surg(?:ery|ical)\s*report"],
    "pathology": [r"patholog|biopsy|histolog"],
    "prescription": [r"prescription|medication\s*list|rx\b"],
    "referral": [r"referral|consult\s*request"],
    "insurance": [r"explanation\s*of\s*benefits|eob|claim|insurance"],
    "mental_health": [r"psychiatr|psycholog|mental\s*health|therapy\s*note"],
}

def auto_categorize_medical(text: str, filename: str = "") -> str:
    """Auto-categorize a medical document based on content."""
    combined = f"{filename}\n{text[:2000]}"
    for category, patterns in MEDICAL_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, re.IGNORECASE):
                return category
    return "other"


def extract_date_from_text(text: str, filename: str = "") -> Optional[str]:
    """Try to extract a date of service from text or filename."""
    # Common date patterns
    patterns = [
        r'(\d{1,2})[/-](\d{1,2})[/-](20\d{2})',  # MM/DD/YYYY
        r'(20\d{2})[/-](\d{1,2})[/-](\d{1,2})',  # YYYY-MM-DD
    ]
    for p in patterns:
        m = re.search(p, filename + " " + text[:1000])
        if m:
            groups = m.groups()
            try:
                if len(groups[0]) == 4:
                    return f"{groups[0]}-{int(groups[1]):02d}-{int(groups[2]):02d}"
                else:
                    return f"{groups[2]}-{int(groups[0]):02d}-{int(groups[1]):02d}"
            except (ValueError, IndexError):
                pass
    return None


# ============================================================
# Embedding client
# ============================================================

class EmbeddingClient:
    """Local BGE-M3 embedding client with batching (delegates to embedding_service)."""

    def __init__(self):
        self.total_tokens = 0

    async def embed_batch(self, texts: List[str]) -> List[List[float]]:
        if not texts:
            return []
        return embed_texts_sync(texts)

    async def close(self):
        pass  # No HTTP client to close


def _vec_literal(embedding: List[float]) -> str:
    return _embedding_vec_literal(embedding)


# ============================================================
# Database operations
# ============================================================

async def file_already_ingested(pool: asyncpg.Pool, source_path: str) -> bool:
    """Check if a file has already been ingested."""
    row = await pool.fetchrow(
        "SELECT id FROM core.documents WHERE source_path = $1", source_path
    )
    return row is not None


async def ensure_patient_exists(conn, name: str) -> int:
    row = await conn.fetchrow(
        "SELECT id FROM medical.patients WHERE name = $1", name
    )
    if row:
        return row["id"]
    row = await conn.fetchrow(
        "INSERT INTO medical.patients (name) VALUES ($1) ON CONFLICT (name) DO NOTHING RETURNING id",
        name
    )
    if row:
        return row["id"]
    row = await conn.fetchrow("SELECT id FROM medical.patients WHERE name = $1", name)
    return row["id"] if row else None


async def ensure_provider_exists(conn, name: str, facility: str = None) -> Optional[int]:
    if not name:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM medical.providers WHERE name = $1", name
    )
    if row:
        return row["id"]
    row = await conn.fetchrow(
        "INSERT INTO medical.providers (name, facility) VALUES ($1, $2) RETURNING id",
        name, facility or ""
    )
    return row["id"] if row else None


async def write_document(pool: asyncpg.Pool, doc_id: str, domain: str,
                         source_path: str, filename: str, document_type: str,
                         title: str, content_hash: str, full_content: str,
                         metadata: dict, chunks: List[str],
                         embeddings: List[List[float]],
                         medical_meta: Optional[dict] = None,
                         enriched_texts: Optional[List[str]] = None):
    """Write document + chunks + enrichment to PostgreSQL."""
    async with pool.acquire() as conn:
        async with conn.transaction():
            # core.documents
            await conn.execute("""
                INSERT INTO core.documents
                    (id, domain, source_path, filename, document_type, title,
                     content_hash, total_chunks, full_content, metadata)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb)
                ON CONFLICT (id) DO NOTHING
            """,
                uuid.UUID(doc_id), domain, source_path,
                filename, document_type, title,
                content_hash, len(chunks),
                full_content[:500000],
                json.dumps(metadata),
            )

            # core.document_chunks
            for i, (chunk_text, emb) in enumerate(zip(chunks, embeddings)):
                chunk_id = hashlib.md5(
                    f"{source_path}:{i}".encode()
                ).hexdigest()
                enriched = enriched_texts[i] if enriched_texts and i < len(enriched_texts) else None
                await conn.execute("""
                    INSERT INTO core.document_chunks
                        (id, document_id, chunk_index, total_chunks,
                         content, embedded_content, embedding, metadata)
                    VALUES ($1, $2, $3, $4, $5, $6, $7::halfvec, $8::jsonb)
                    ON CONFLICT (id) DO NOTHING
                """,
                    chunk_id, uuid.UUID(doc_id),
                    i, len(chunks),
                    chunk_text, enriched, _vec_literal(emb),
                    json.dumps({}),
                )

            # Medical enrichment
            if domain == "medical" and medical_meta:
                patient_id = None
                if medical_meta.get("patient_name"):
                    patient_id = await ensure_patient_exists(
                        conn, medical_meta["patient_name"]
                    )

                provider_id = None
                if medical_meta.get("provider"):
                    provider_id = await ensure_provider_exists(
                        conn, medical_meta["provider"], medical_meta.get("facility")
                    )

                dos = medical_meta.get("date_of_service")
                dos_date = None
                if dos:
                    try:
                        dos_date = datetime.strptime(dos, "%Y-%m-%d").date()
                    except (ValueError, TypeError):
                        pass

                await conn.execute("""
                    INSERT INTO medical.record_metadata
                        (document_id, patient_id, provider_id, record_type,
                         date_of_service, facility, metadata)
                    VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb)
                    ON CONFLICT (document_id) DO NOTHING
                """,
                    uuid.UUID(doc_id),
                    patient_id, provider_id,
                    medical_meta.get("record_type", "other"),
                    dos_date,
                    medical_meta.get("facility", ""),
                    json.dumps({"source": "dropbox", "ingested_at": datetime.now().isoformat()}),
                )

            # Legal enrichment — link to cases by case number
            if domain == "legal":
                case_numbers = extract_case_numbers(full_content)
                for case_num in case_numbers:
                    row = await conn.fetchrow(
                        "SELECT id FROM legal.cases WHERE case_number = $1",
                        case_num
                    )
                    if row:
                        await conn.execute("""
                            INSERT INTO legal.case_documents (case_id, document_id)
                            VALUES ($1, $2) ON CONFLICT DO NOTHING
                        """, row["id"], uuid.UUID(doc_id))


# Case number extraction
STATE_CASE_RE = re.compile(r'(\d{2})-(\d)-(\d{5})-(\d{2})')
FEDERAL_CASE_RE = re.compile(
    r'(\d{1,2})[-:]?(\d{2})[-:]?cv[-:]?(\d{5})[-:]?(\w{2,4})?', re.IGNORECASE
)

def extract_case_numbers(text: str) -> List[str]:
    found = set()
    for m in FEDERAL_CASE_RE.finditer(text[:5000]):
        court, year, seq, judge = m.groups()
        normalized = f"{court}{year}CV{seq}"
        if judge:
            normalized += judge.upper()
        found.add(normalized.upper())
    for m in STATE_CASE_RE.finditer(text[:5000]):
        found.add(m.group(0))
    return list(found)


# ============================================================
# File processor
# ============================================================

async def process_file(pool: asyncpg.Pool, embedder: EmbeddingClient,
                       file_path: Path, domain: str) -> bool:
    """Process a single dropped file end-to-end. Returns True on success."""
    source_path = str(file_path)
    filename = file_path.name

    # Skip if already ingested
    if await file_already_ingested(pool, source_path):
        log.info(f"  Already ingested: {filename}")
        return True

    log.info(f"  Extracting text from {filename} ({file_path.stat().st_size:,} bytes)")
    text = extract_text(file_path)

    if not text or len(text.strip()) < 20:
        log.warning(f"  No meaningful text extracted from {filename}")
        return False

    # Chunk
    chunker = TextChunker()
    chunks = chunker.split(text)
    if not chunks:
        log.warning(f"  No chunks produced for {filename}")
        return False

    log.info(f"  {len(chunks)} chunks, embedding...")

    # Contextual Retrieval: generate context and enrich chunks
    context = generate_context_sync(
        title=file_path.stem.replace("_", " ").replace("-", " "),
        domain=domain,
        document_type="document",
        content_preview=text[:3000],
        case_number=None,
    )
    enriched_texts = enrich_chunks(context, chunks)

    # Embed enriched texts (batch in groups of 2000)
    all_embeddings = []
    for i in range(0, len(enriched_texts), 2000):
        batch = enriched_texts[i:i+2000]
        try:
            embs = await embedder.embed_batch(batch)
            all_embeddings.extend(embs)
        except Exception as e:
            log.error(f"  Embedding failed for {filename}: {e}")
            # Zero-fill fallback
            all_embeddings.extend([[0.0] * EMBEDDING_DIMENSIONS] * len(batch))

    # Determine document type
    ext = file_path.suffix.lower()
    if domain == "medical":
        doc_type = auto_categorize_medical(text, filename)
        medical_meta = {
            "patient_name": "William Nelson",
            "record_type": doc_type,
            "date_of_service": extract_date_from_text(text, filename),
            "provider": "",
            "facility": "UW Medicine" if "uw" in filename.lower() or "university of washington" in text[:2000].lower() else "",
        }
    elif domain == "legal":
        doc_type = "court_filing" if ext == ".pdf" else (
            "email" if ext == ".eml" else "legal_document"
        )
        medical_meta = None
    else:
        doc_type = "document"
        medical_meta = None

    doc_id = str(uuid.uuid4())
    content_hash = hashlib.sha256(text.encode()).hexdigest()[:16]
    title = file_path.stem.replace("_", " ").replace("-", " ")

    log.info(f"  Writing to DB: domain={domain}, type={doc_type}, chunks={len(chunks)}")

    try:
        await write_document(
            pool, doc_id, domain, source_path, filename,
            doc_type, title, content_hash, text,
            {"file_type": ext.lstrip("."), "source": "samba_dropbox",
             "dropped_at": datetime.now().isoformat()},
            chunks, all_embeddings, medical_meta,
            enriched_texts=enriched_texts,
        )

        # Send push notification about new document
        try:
            import httpx as _httpx
            _ntfy_topic = os.getenv("NTFY_TOPIC", "")
            if _ntfy_topic:
                async with _httpx.AsyncClient(timeout=5) as _client:
                    await _client.post(
                        f"https://ntfy.sh/{_ntfy_topic}",
                        headers={
                            "Title": f"📄 New {domain} document ingested",
                            "Priority": "3",
                            "Tags": "page_facing_up",
                        },
                        content=f"{filename}\nType: {doc_type}\nChunks: {len(chunks)}\nNow searchable in Athena",
                    )
        except Exception:
            pass  # Notification is best-effort

        return True
    except Exception as e:
        log.error(f"  DB write failed for {filename}: {e}")
        return False


# ============================================================
# Main watcher loop
# ============================================================

def get_domain_for_path(file_path: Path) -> str:
    """Determine domain based on the subdirectory."""
    rel = file_path.relative_to(DROPBOX_DIR)
    parts = rel.parts
    if parts and parts[0] == "medical":
        return "medical"
    elif parts and parts[0] == "legal":
        return "legal"
    return "general"


def find_new_files() -> List[Path]:
    """Scan dropbox for unprocessed files."""
    files = []
    for ext in SUPPORTED_EXTENSIONS:
        files.extend(DROPBOX_DIR.rglob(f"*{ext}"))
    # Exclude hidden dirs (.processed, .failed)
    return [f for f in files if not any(p.startswith(".") for p in f.relative_to(DROPBOX_DIR).parts)]


def move_to_processed(file_path: Path, success: bool):
    """Move file to .processed/ or .failed/ preserving subdirectory."""
    domain = get_domain_for_path(file_path)
    target_dir = (PROCESSED_DIR if success else FAILED_DIR) / domain
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / file_path.name

    # Handle name collisions
    if target.exists():
        stem = file_path.stem
        suffix = file_path.suffix
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = target_dir / f"{stem}_{ts}{suffix}"

    shutil.move(str(file_path), str(target))
    log.info(f"  Moved to {target.relative_to(DROPBOX_DIR)}")


async def process_pending(pool: asyncpg.Pool, embedder: EmbeddingClient) -> int:
    """Find and process all pending files. Returns count processed."""
    files = find_new_files()
    if not files:
        return 0

    log.info(f"Found {len(files)} file(s) to process")
    processed = 0

    for file_path in sorted(files):
        domain = get_domain_for_path(file_path)
        log.info(f"Processing: {file_path.relative_to(DROPBOX_DIR)} → {domain}")

        try:
            success = await process_file(pool, embedder, file_path, domain)
            move_to_processed(file_path, success)
            if success:
                processed += 1
        except Exception as e:
            log.error(f"Unhandled error processing {file_path.name}: {e}")
            move_to_processed(file_path, False)

    return processed


async def watch_loop():
    """Main daemon loop using inotify to watch for new files."""
    # Ensure directories exist
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    FAILED_DIR.mkdir(parents=True, exist_ok=True)
    for sub in ("medical", "legal", "general"):
        (DROPBOX_DIR / sub).mkdir(parents=True, exist_ok=True)

    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=5)
    embedder = EmbeddingClient()

    log.info("=" * 60)
    log.info("WDWS Dropbox Watcher started")
    log.info(f"  Watching: {DROPBOX_DIR}")
    log.info(f"  Subdirs:  medical/ legal/ general/")
    log.info(f"  Database: {DATABASE_URL.split('@')[1] if '@' in DATABASE_URL else DATABASE_URL}")
    log.info("=" * 60)

    # Process any files that were dropped while we were offline
    initial = await process_pending(pool, embedder)
    if initial:
        log.info(f"Processed {initial} backlog file(s)")

    # Set up inotify watches
    inotify = INotify()
    watch_flags = (
        inotify_flags.CLOSE_WRITE |
        inotify_flags.MOVED_TO
    )

    # Watch each subdirectory
    watch_dirs = {}
    for sub in ("medical", "legal", "general"):
        d = DROPBOX_DIR / sub
        wd = inotify.add_watch(str(d), watch_flags)
        watch_dirs[wd] = d
        log.info(f"  Watching: {d}")

    # Also watch root for files dropped there
    wd = inotify.add_watch(str(DROPBOX_DIR), watch_flags)
    watch_dirs[wd] = DROPBOX_DIR

    pending_files: Dict[Path, float] = {}

    log.info("Waiting for files...")

    while True:
        # Check inotify with a 2-second timeout
        events = inotify.read(timeout=2000)

        for event in events:
            if event.name:
                watch_dir = watch_dirs.get(event.wd, DROPBOX_DIR)
                file_path = watch_dir / event.name

                if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_EXTENSIONS:
                    # Don't process hidden files
                    if not event.name.startswith("."):
                        pending_files[file_path] = time.time()
                        log.info(f"Detected: {file_path.relative_to(DROPBOX_DIR)}")

        # Process files that have settled (no writes for SETTLE_SECONDS)
        now = time.time()
        ready = [
            fp for fp, ts in pending_files.items()
            if now - ts >= SETTLE_SECONDS and fp.exists()
        ]

        if ready:
            for file_path in ready:
                del pending_files[file_path]
                domain = get_domain_for_path(file_path)
                log.info(f"Processing: {file_path.relative_to(DROPBOX_DIR)} → {domain}")
                try:
                    success = await process_file(pool, embedder, file_path, domain)
                    move_to_processed(file_path, success)
                    if success:
                        log.info(f"✔ Ingested: {file_path.name}")
                except Exception as e:
                    log.error(f"✗ Failed: {file_path.name}: {e}")
                    move_to_processed(file_path, False)

        # Yield to event loop
        await asyncio.sleep(0.1)


async def run_once():
    """Process pending files and exit."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    FAILED_DIR.mkdir(parents=True, exist_ok=True)

    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=5)
    embedder = EmbeddingClient()

    count = await process_pending(pool, embedder)
    log.info(f"Done. Processed {count} file(s).")

    await embedder.close()
    await pool.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="WDWS Dropbox Watcher")
    parser.add_argument("--once", action="store_true", help="Process pending files and exit")
    args = parser.parse_args()

    if args.once:
        asyncio.run(run_once())
    else:
        asyncio.run(watch_loop())
